from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from datetime import date, timedelta, datetime
from app.models import Schedule, Employee, Task
from app import db
from sqlalchemy.exc import IntegrityError
import io

schedules_bp = Blueprint('schedules', __name__)

DAYS_VI = {0: 'Thứ Hai', 1: 'Thứ Ba', 2: 'Thứ Tư',
           3: 'Thứ Năm', 4: 'Thứ Sáu', 5: 'Thứ Bảy'}
SHIFTS = ['morning', 'afternoon']
SHIFT_LABELS = {'morning': 'Ca Sáng\n07:30–11:30', 'afternoon': 'Ca Chiều\n13:00–17:00'}


def get_week_range(ref_date: date):
    """Return Monday and Saturday of the week containing ref_date."""
    weekday = ref_date.weekday()  # Mon=0, Sun=6
    monday = ref_date - timedelta(days=weekday)
    saturday = monday + timedelta(days=5)
    return monday, saturday


@schedules_bp.route('/')
def weekly_view():
    ref_str = request.args.get('week')
    if ref_str:
        try:
            ref_date = datetime.strptime(ref_str, '%Y-%m-%d').date()
        except ValueError:
            ref_date = date.today()
    else:
        ref_date = date.today()

    monday, saturday = get_week_range(ref_date)
    prev_week = (monday - timedelta(days=7)).strftime('%Y-%m-%d')
    next_week = (monday + timedelta(days=7)).strftime('%Y-%m-%d')

    # Build list of working dates (Mon–Sat)
    work_dates = [monday + timedelta(days=i) for i in range(6)]

    # Fetch all schedules in this week
    schedules = Schedule.query.filter(
        Schedule.work_date >= monday,
        Schedule.work_date <= saturday
    ).all()

    # Build matrix: grid[date][shift] = list of schedules
    grid = {}
    for d in work_dates:
        grid[d] = {'morning': [], 'afternoon': []}
    for s in schedules:
        if s.work_date in grid and s.shift in grid[s.work_date]:
            grid[s.work_date][s.shift].append(s)

    return render_template('schedules/weekly.html',
                           grid=grid,
                           work_dates=work_dates,
                           days_vi=DAYS_VI,
                           shift_labels=SHIFT_LABELS,
                           prev_week=prev_week,
                           next_week=next_week,
                           current_week_start=monday,
                           today=date.today())


@schedules_bp.route('/assign', methods=['GET', 'POST'])
def assign():
    employees = Employee.query.order_by(Employee.full_name).all()
    tasks = Task.query.order_by(Task.task_name).all()

    if request.method == 'POST':
        emp_id = request.form.get('employee_id', type=int)
        task_id = request.form.get('task_id', type=int)
        work_date_str = request.form.get('work_date', '').strip()
        shift = request.form.get('shift', '').strip()
        note = request.form.get('note', '').strip()

        # Validation
        errors = []
        if not emp_id:
            errors.append('Vui lòng chọn nhân viên.')
        if not task_id:
            errors.append('Vui lòng chọn công việc.')
        if not work_date_str:
            errors.append('Vui lòng chọn ngày làm việc.')
        if shift not in SHIFTS:
            errors.append('Ca làm việc không hợp lệ.')

        if not errors:
            try:
                work_date = datetime.strptime(work_date_str, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Định dạng ngày không hợp lệ.')

        if not errors:
            # Check valid weekday (Mon–Sat, weekday 0–5)
            if work_date.weekday() > 5:
                errors.append('Chỉ được phân công từ Thứ Hai đến Thứ Bảy.')

            # Check duplicate shift for this employee
            existing = Schedule.query.filter_by(
                employee_id=emp_id,
                work_date=work_date,
                shift=shift
            ).first()
            if existing:
                emp = Employee.query.get(emp_id)
                errors.append(
                    f'Nhân viên {emp.full_name} đã có lịch vào {SHIFT_LABELS[shift].replace(chr(10), " ")} '
                    f'ngày {work_date.strftime("%d/%m/%Y")}!'
                )

        if errors:
            for e in errors:
                flash(e, 'danger')
        else:
            try:
                schedule = Schedule(
                    employee_id=emp_id,
                    task_id=task_id,
                    work_date=work_date,
                    shift=shift,
                    note=note,
                )
                db.session.add(schedule)
                db.session.commit()
                flash('Phân công công việc thành công!', 'success')
                return redirect(url_for('schedules.weekly_view',
                                        week=work_date.strftime('%Y-%m-%d')))
            except IntegrityError:
                db.session.rollback()
                flash('Lỗi: Nhân viên đã bị trùng ca làm việc!', 'danger')
            except Exception as e:
                db.session.rollback()
                flash(f'Lỗi hệ thống: {str(e)}', 'danger')

    return render_template('schedules/assign.html',
                           employees=employees,
                           tasks=tasks,
                           shifts=SHIFTS,
                           shift_labels=SHIFT_LABELS,
                           today_str=date.today().strftime('%Y-%m-%d'))


@schedules_bp.route('/<int:schedule_id>/delete', methods=['POST'])
def delete_schedule(schedule_id):
    s = Schedule.query.get_or_404(schedule_id)
    week = s.work_date.strftime('%Y-%m-%d')
    try:
        db.session.delete(s)
        db.session.commit()
        flash('Đã xóa phân công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {str(e)}', 'danger')
    return redirect(url_for('schedules.weekly_view', week=week))


@schedules_bp.route('/export')
def export_excel():
    """Export weekly schedule to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl chưa được cài đặt!', 'danger')
        return redirect(url_for('schedules.weekly_view'))

    ref_str = request.args.get('week')
    if ref_str:
        try:
            ref_date = datetime.strptime(ref_str, '%Y-%m-%d').date()
        except ValueError:
            ref_date = date.today()
    else:
        ref_date = date.today()

    monday, saturday = get_week_range(ref_date)
    work_dates = [monday + timedelta(days=i) for i in range(6)]

    schedules = Schedule.query.filter(
        Schedule.work_date >= monday,
        Schedule.work_date <= saturday
    ).all()

    grid = {}
    for d in work_dates:
        grid[d] = {'morning': [], 'afternoon': []}
    for s in schedules:
        if s.work_date in grid and s.shift in grid[s.work_date]:
            grid[s.work_date][s.shift].append(s)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Lịch tuần {monday.strftime('%d/%m')}-{saturday.strftime('%d/%m/%Y')}"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    shift_fill = PatternFill("solid", fgColor="2D6A9F")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"LỊCH LÀM VIỆC TUẦN {monday.strftime('%d/%m')} - {saturday.strftime('%d/%m/%Y')}"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal='center')

    # Header row
    ws['A2'] = 'Ca / Ngày'
    ws['A2'].font = white_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = Alignment(horizontal='center')

    col = 2
    for d in work_dates:
        ws.cell(row=2, column=col).value = f"{DAYS_VI[d.weekday()]}\n{d.strftime('%d/%m/%Y')}"
        ws.cell(row=2, column=col).font = white_font
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 22
        col += 1

    ws.column_dimensions['A'].width = 20
    ws.row_dimensions[2].height = 35

    row = 3
    for shift in SHIFTS:
        ws.cell(row=row, column=1).value = SHIFT_LABELS[shift].replace('\n', ' ')
        ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=1).fill = shift_fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=1).border = thin_border

        col = 2
        for d in work_dates:
            entries = grid[d][shift]
            cell = ws.cell(row=row, column=col)
            if entries:
                lines = [f"{s.employee.full_name} - {s.task.task_name}" for s in entries]
                cell.value = '\n'.join(lines)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.value = ''
            cell.border = thin_border
            col += 1

        ws.row_dimensions[row].height = 50
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"lich_tuan_{monday.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
